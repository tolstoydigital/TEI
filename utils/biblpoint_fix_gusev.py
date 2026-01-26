import os
import re
import time
from pathlib import Path
from typing import Optional, Dict

from lxml import etree
import logging
from csv import DictWriter
from openpyxl import load_workbook, Workbook
import subprocess


def setup_logger(name: str = None, level=logging.INFO, logfile: Path | None = None):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.handlers.clear()

    formatter = logging.Formatter("%(levelname)s: %(message)s")  # Без даты

    # Вывод в консоль
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # Если указан файл — пишем ещё и туда
    if logfile:
        file_handler = logging.FileHandler(logfile, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


# Абсолютный путь к текущему файлу
BASE_DIR = Path(__file__).resolve().parent
TOP_DIR = BASE_DIR.parent.parent

bibllist_bio_path = BASE_DIR.parent / "reference" / "bibllist_bio.xml"
bibllist_bio_path = bibllist_bio_path.resolve()
bibllist_bio_short_path = bibllist_bio_path.relative_to(TOP_DIR)

SD = "-"
LD = "—"
BIO_C = "С."
pattern = r"(?<=\b\d{4})\s*[-–—]\s*(?=\d{4}\b)"


logger = setup_logger(__name__, logfile=Path("biblpoint_change.log"))


def get_title_and_bibl_scope(
    xml_path: Path, save_to: Optional[Path] = None
) -> Dict[str, str]:
    short_path = xml_path.relative_to(TOP_DIR)

    tree = etree.parse(xml_path)
    root = tree.getroot()

    ns = root.nsmap.get(None)
    nsmap = (
        {"tei": ns, "xml": "http://www.w3.org/XML/1998/namespace"}
        if ns
        else {"xml": "http://www.w3.org/XML/1998/namespace"}
    )

    for title_el in root.findall(
        ".//tei:title[@type='bibl']" if ns else ".//title[@type='bibl']",
        namespaces=nsmap,
    ):
        # original_title = title_el.text
        if title_el.text is None:
            # continue  # защита от пустых тегов
            source_old_title_text = ""
            old_title_text = ""
        else:
            source_old_title_text = title_el.text.strip()
            old_title_text = (
                source_old_title_text  # replace_title(source_old_title_text)
            )

        bibl_scope_text = ""
        for bibl_scope_el in root.findall(
            ".//tei:biblScope[@unit='page']" if ns else ".//biblScope[@unit='page']",
            namespaces=nsmap,
        ):
            bibl_scope_text = bibl_scope_el.text.strip()

        return {"bibl_scope": bibl_scope_text, "title": old_title_text}

    return None


def get_bibllist_bio_title_and_id(filename: Path) -> Optional[Dict[str, str]]:

    xml_id = filename.stem
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }
    xpath = f".//tei:relatedItem[tei:ref[@xml:id=\"{xml_id}\"]]/tei:title[@type='bibl']"
    titles = bib_root.xpath(xpath, namespaces=ns)
    for t in titles:
        # logger.info(f"====================================== {xml_id}")
        # logger.info(f"BIBLIO <<< {t.text}")
        # logger.info(f"BIBLIO >>> {new_title}")
        # t.text = new_title
        bib_changes = {"bibl_original_title": t.text, "xml_id": xml_id}
        # bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
        return bib_changes
    return None


def fix_in_bibllist_bio_case_2(
    document_id: str,
    source_list_item_id: str,
    source_title_bibl: str,
    bibl_segment_pages: Optional[str],
):
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    new_biblpoint = f"{source_title_bibl} {bibl_segment_pages if bibl_segment_pages else ''}".strip()
    if new_biblpoint[-1] != ".":
        new_biblpoint = new_biblpoint + "."

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }
    # root.xpath('//ns:ref[@target="slovar"]', namespaces={'ns': XMLNS})
    xpath = (
        f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]/'
        f'tei:relation[@type="source" and @ref="{source_list_item_id}"]'
    )  # /tei:biblpoint
    xpath_empty = (
        f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]/'
        f"tei:relation[@type='source' and @ref='EMPTY' and @source='tolstoy-bio']"
    )
    xpath_related_item = f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]'
    relations = bib_root.xpath(xpath, namespaces=ns)
    logger.info(f"======= CASE_2 id={document_id} ref='{source_list_item_id}'")
    logger.info(f"      {xpath=}")
    if relations:
        for rel in relations:
            bp_xpath = f".//tei:biblpoint"
            bps = rel.xpath(bp_xpath, namespaces=ns)
            if bps:
                for bp in bps:
                    if bp.text == new_biblpoint:
                        logger.info(
                            f"      biblpoint is already exists in {document_id}"
                        )
                    else:
                        logger.info(
                            f"       biblpoint changed '{bp.text}' -> '{new_biblpoint}'"
                        )
                        bp.text = new_biblpoint
                        # print(etree.tostring(bp, pretty_print=True, encoding="unicode"))
            else:
                new_node = etree.SubElement(rel, "biblpoint")
                new_node.text = new_biblpoint
                logger.info(
                    f'      created new node {etree.tostring(new_node, pretty_print=True, encoding="unicode")}'
                )

    else:  # thera is to any relation with @ref="{source_list_item_id}"
        logger.warning(
            f"       node with {source_list_item_id} does not found, trying to find EMPTY node"
        )
        relations_2 = bib_root.xpath(xpath_empty, namespaces=ns)

        if relations_2:
            for rel in relations_2:
                bp_xpath = f".//tei:biblpoint"
                bps = rel.xpath(bp_xpath, namespaces=ns)
                if bps:
                    for bp in bps:
                        if bp.text == new_biblpoint:
                            logger.info(
                                f"      biblpoint is already exists in {document_id}"
                            )
                        else:
                            logger.info(
                                f"       biblpoint changed '{bp.text}' -> '{new_biblpoint}'"
                            )
                            bp.text = new_biblpoint
                            # print(etree.tostring(bp, pretty_print=True, encoding="unicode"))
                else:
                    new_node = etree.SubElement(rel, "biblpoint")
                    new_node.text = new_biblpoint
                    logger.info(
                        f'      created new node {etree.tostring(new_node, pretty_print=True, encoding="unicode")}'
                    )

        else:
            logger.info(
                f'      NO relation ref="EMPTY" source="tolstoy-bio" type="source"'
            )
            rel_it_nodes = bib_root.xpath(xpath_related_item, namespaces=ns)
            if rel_it_nodes:
                for _rel in rel_it_nodes:

                    new_rel_node = etree.SubElement(
                        _rel,
                        "relation",
                        type="source",
                        ref=source_list_item_id,
                        source="tolstoy-bio",
                    )
                    new_node = etree.SubElement(new_rel_node, "biblpoint")
                    new_node.text = new_biblpoint
                    logger.info(
                        f'      created new node {etree.tostring(new_rel_node, pretty_print=True, encoding="unicode")}'
                    )
            else:
                logger.error(f"      NO relatedItem")
    bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
    return


def fix_in_bibllist_bio_case_3(
    document_id: str,
    td_source_item_id: str,
    source: str,
    title_main_td_source_item: str,
    url_td_source_item_id_for_bibl_point: str,
):
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }
    td_source_item_id = td_source_item_id if td_source_item_id else ""
    url_td_source_item_id_for_bibl_point = url_td_source_item_id_for_bibl_point if url_td_source_item_id_for_bibl_point else ""
    title_main_td_source_item = title_main_td_source_item if title_main_td_source_item else ""

    # # root.xpath('//ns:ref[@target="slovar"]', namespaces={'ns': XMLNS})
    # xpath = (f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]/'
    #          f'tei:relation[@type="source" and @ref="{source_list_item_id}"]')  # /tei:biblpoint
    # xpath_empty = (
    #     f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]/'
    #     f"tei:relation[@type='source' and @ref='EMPTY' and @source='tolstoy-bio']"
    # )
    xpath_related_item = f'.//tei:relatedItem[tei:ref[@xml:id="{document_id}"]]'
    rel_it_nodes = bib_root.xpath(xpath_related_item, namespaces=ns)
    logger.info(f"======= CASE_3 id={document_id}")
    logger.info(f"      {xpath_related_item=}")
    if rel_it_nodes:
        for _rel in rel_it_nodes:
            logger.info(
                f'      {etree.tostring(_rel, pretty_print=True, encoding="unicode")}'
            )
            # все relation у родителя
            relations = _rel.xpath(".//tei:relation", namespaces=ns)
            if relations:
                last_relation = relations[-1]
                insert_index = _rel.index(last_relation) + 1
            else:
                # если relation вообще нет — вставляем в конец
                insert_index = len(_rel)
            added = False
            match source:
                case "slovo-tolstogo":
                    new_rel_node = etree.Element(
                        _tag="relation",
                        ref=td_source_item_id,
                        type="source",
                        source=source,
                    )
                    new_biblpoint_node = etree.SubElement(new_rel_node, "biblpoint")
                    new_biblpoint_node.text = title_main_td_source_item

                    new_url_node = etree.SubElement(new_rel_node, "url")
                    new_url_node.text = url_td_source_item_id_for_bibl_point

                    _rel.insert(insert_index, new_rel_node)
                    added = True
                case "tolstoy-bio":
                    new_rel_node = etree.Element(
                        _tag="relation",
                        ref=td_source_item_id,
                        type="source",
                        source=source,
                    )
                    _rel.insert(insert_index, new_rel_node)
                    added = True
                case _:  # HZ
                    pass
            if added:
                logger.info(
                    f'      added node {etree.tostring(new_rel_node, pretty_print=True, encoding="unicode")}'
                )
            else:
                logger.warning(f"      nothing was added")

    else:
        logger.info(f"      NO relatedItem")

    bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
    return

def fix_in_bibllist_bio_case_4():
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }

    xpath = (f'.//tei:relatedItem[tei:ref[@xml:id]]/'
             f'tei:relation[@source="tolstoy-bio" and @ref!="EMPTY"]')  # /tei:biblpoint

    xp_title = f".//tei:title[@type='main']"
    xp_date = f".//tei:date[@type='editor']"

    rel_it_nodes = bib_root.xpath(xpath, namespaces=ns)
    logger.info(f"======= CASE_4")
    if rel_it_nodes:
        for _rel in rel_it_nodes:
            logger.info(
                f'      ------ {etree.tostring(_rel, pretty_print=True, encoding="unicode")}'
            )
            ref_value = _rel.attrib.get("ref")
            xpath_2 = f'.//tei:relatedItem[tei:ref[@xml:id="{ref_value}"]]'
            main_rel_nodes = bib_root.xpath(xpath_2, namespaces=ns)
            if main_rel_nodes:
                for _rel_main in main_rel_nodes:
                    titles = _rel_main.xpath(xp_title, namespaces=ns)
                    if titles:
                        title = titles[0].text.strip()
                    else:
                        title = ""

                    dates = _rel_main.xpath(xp_date, namespaces=ns)
                    if dates:
                        date = dates[0].text.strip()
                    else:
                        date = ""
                    if title and date:
                        # check if biblpoint is in relation
                        bibls = _rel.xpath('.//tei:biblpoint', namespaces=ns)
                        if bibls:
                            logger.error(f'      relation already has {etree.tostring(bibls[0], pretty_print=True, encoding="unicode")}')
                        else:
                            new_biblpoint_node = etree.SubElement(_rel, "biblpoint")
                            title = title.rstrip('.')
                            new_biblpoint_node.text = f"{title}, {date}."
                            logger.info(
                                f'      added node {etree.tostring(new_biblpoint_node, pretty_print=True, encoding="unicode")}'
                            )
            else:
                logger.warning(f"      NO title and date")
    else:
        logger.info(f"      NO relatedItem")

    bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
    return


def main():
    start = time.perf_counter()

    ref_path = (
        BASE_DIR.parent
        / "utils"
        / "doc"
        / "post-evaluation-table - True_False & pages Review_Edits_02.xlsx"
    )
    ref_path = ref_path.resolve()

    export_csv_path = BASE_DIR.parent / "utils"

    try:
        workbook = load_workbook(ref_path)
    except FileNotFoundError:
        print(f"Файл {ref_path} не найден.")
        return

    # sheet = workbook.active
    sheet = workbook["Sheet1"]

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(
        min_row=2,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column,
        values_only=True,
    ):
        if not any(row):
            continue
        data = dict(zip(headers, row))
        is_linked = data.get("is_linked", False)
        is_manually_linked = data.get("is_manually_linked", False)
        document_id = data.get("document_id", "")
        source_list_item_id = data.get("source_list_item_id", "")  # K
        source_title_bibl = data.get("source_title_bibl", "")
        bibl_segment_pages = data.get("bibl_segment_pages", "")
        td_source_item_id = data.get("TD_source_item_ID", "")  # N
        source = data.get("source", "")  # T
        title_main_td_source_item = data.get("title_main_TD_source_item", "")  # M
        url_td_source_item_id_for_bibl_point = data.get(
            "url_TD_source_item_ID_for_bibl_point", ""
        )

        if not any((is_linked, is_manually_linked)):
            continue
        if is_linked:
            fix_in_bibllist_bio_case_2(
                document_id, source_list_item_id, source_title_bibl, bibl_segment_pages
            )
        if is_manually_linked:
            fix_in_bibllist_bio_case_3(
                document_id,
                td_source_item_id,
                source,
                title_main_td_source_item,
                url_td_source_item_id_for_bibl_point,
            )
        # print(data)

    # global change over entire bibllist_bio
    fix_in_bibllist_bio_case_4()

    elapsed = time.perf_counter() - start
    logger.info(f"Total time: {elapsed:.2f} seconds")


if __name__ == "__main__":
    main()
