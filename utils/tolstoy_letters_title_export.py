import os
import re
import time
from pathlib import Path
from typing import Optional, Dict

from lxml import etree
import logging
from csv import DictWriter
from openpyxl import load_workbook, Workbook
import subprocess

def setup_logger(name: str = None, level=logging.INFO, logfile: Path | None = None):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.handlers.clear()

    formatter = logging.Formatter("%(levelname)s: %(message)s")  # Без даты

    # Вывод в консоль
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # Если указан файл — пишем ещё и туда
    if logfile:
        file_handler = logging.FileHandler(logfile, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


# Абсолютный путь к текущему файлу
BASE_DIR = Path(__file__).resolve().parent
TOP_DIR = BASE_DIR.parent.parent

bibllist_bio_path = BASE_DIR.parent / "reference" / "bibllist_bio.xml"
bibllist_bio_path = bibllist_bio_path.resolve()
bibllist_bio_short_path = bibllist_bio_path.relative_to(TOP_DIR)

SD = "-"
LD = "—"
BIO_C = "С."
pattern = r"(?<=\b\d{4})\s*[-–—]\s*(?=\d{4}\b)"

TITLE_MAP = {
    "Толстая С. А. Дневники, 1862—1900 // Толстая С. А. Дневники: В 2 т. — М.: Худож. лит., 1978 . — Т. 1: 1862—1900. — С. 35—468. — (Сер. лит. мемуаров).": "Толстая С. А. Дневники, 1862—1900 // Толстая С. А. Дневники: В 2 т. М.: Худож. лит., 1978 . Т. 1. С.",
    "Толстая С. А. Дневники, 1901—1910 // Толстая С. А. Дневники: В 2 т. — М.: Худож. лит., 1978 . — Т. 2: 1901—1910. — С. 5—226. — (Сер. лит. мемуаров).": "Толстая С. А. Дневники, 1901—1910 // Толстая С. А. Дневники: В 2 т.  М.: Худож. лит., 1978. Т. 2. С.",
}

logger = setup_logger(__name__, logfile=Path("mylog_tolstaya_diaries.log"))


def replace_title(title: str) -> str:
    return TITLE_MAP.get(title, title)


def iter_xml_files(root_dir):
    root_dir = Path(root_dir)
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            if filename.lower().endswith(".xml"):
                yield Path(dirpath) / filename


def get_title_and_bibl_scope(xml_path: Path, save_to: Optional[Path] = None) -> Dict[str, str]:
    short_path = xml_path.relative_to(TOP_DIR)

    tree = etree.parse(xml_path)
    root = tree.getroot()

    ns = root.nsmap.get(None)
    nsmap = (
        {"tei": ns, "xml": "http://www.w3.org/XML/1998/namespace"}
        if ns
        else {"xml": "http://www.w3.org/XML/1998/namespace"}
    )

    for title_el in root.findall(
        ".//tei:title[@type='bibl']" if ns else ".//title[@type='bibl']",
        namespaces=nsmap,
    ):
        # original_title = title_el.text
        if title_el.text is None:
            # continue  # защита от пустых тегов
            source_old_title_text = ""
            old_title_text = ""
        else:
            source_old_title_text = title_el.text.strip()
            old_title_text = source_old_title_text  # replace_title(source_old_title_text)

        bibl_scope_text = ""
        for bibl_scope_el in root.findall(
                ".//tei:biblScope[@unit='page']" if ns else ".//biblScope[@unit='page']",
                namespaces=nsmap,
        ):
            bibl_scope_text = bibl_scope_el.text.strip()

        return {"bibl_scope": bibl_scope_text, "title": old_title_text}

    return None


def get_bibllist_bio_title_and_id(
    filename: Path
) -> Optional[Dict[str, str]]:


    xml_id = filename.stem
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }
    xpath = f".//tei:relatedItem[tei:ref[@xml:id=\"{xml_id}\"]]/tei:title[@type='bibl']"
    titles = bib_root.xpath(xpath, namespaces=ns)
    for t in titles:
        # logger.info(f"====================================== {xml_id}")
        # logger.info(f"BIBLIO <<< {t.text}")
        # logger.info(f"BIBLIO >>> {new_title}")
        # t.text = new_title
        bib_changes = {"bibl_original_title": t.text, "xml_id": xml_id}
        # bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
        return bib_changes
    return None



def git_short_sha() -> str:
    return subprocess.check_output(
        ["git", "rev-parse", "--short", "HEAD"],
        stderr=subprocess.DEVNULL,
        text=True,
    ).strip() or ''


def find_file_recursively(search_directory: str, filename: str) -> Optional[str]:
    """
    Ищет файл filename в папке search_directory и всех её подпапках.

    Args:
        search_directory (str): Путь к папке, откуда начинать поиск.
        filename (str): Точное имя файла (например, 'data.xml').

    Returns:
        Optional[str]: Полный абсолютный путь к файлу (str) или None, если файл не найден.
    """
    path_obj = Path(search_directory)

    # Проверяем, существует ли вообще папка поиска
    if not path_obj.exists():
        return None

    try:
        # rglob('*') ищет рекурсивно.
        # Если мы передаем конкретное имя файла в rglob, он будет искать именно его.
        # next() берет первое совпадение.
        found_file = next(path_obj.rglob(filename))

        # .resolve() возвращает абсолютный путь (например, /home/user/projects/...)
        return found_file.resolve()

    except StopIteration:
        # Если генератор пуст (файл не найден), next() вызовет StopIteration
        return None



def main():
    start = time.perf_counter()
    new_file_prefix = "tolstoy-letters_"

    last_git_hash = git_short_sha()

    stage_path = (
        BASE_DIR.parent / "texts" / "letters"
    )
    stage_path = stage_path.resolve()

    stage_front_path = (
        BASE_DIR.parent / "texts_front" / "letters"
    )
    stage_front_path = stage_front_path.resolve()
    export_csv_path = BASE_DIR.parent / "utils"
    stage_front_path.resolve()

    template = 'Схема импорта и экспорта Библиографии.xlsx'

    try:
        workbook = load_workbook(template)
    except FileNotFoundError:
        print(f"Файл {template} не найден.")
        return

    # sheet = workbook.active
    sheet = workbook['Лист1']

    next_row = 3  # sheet.max_row + 1

    for i, xml_file in enumerate(iter_xml_files(stage_path)):
        short_path = xml_file.relative_to(TOP_DIR)
        fn = xml_file.stem
        # if not fn.startswith(new_file_prefix):
        #     continue

        # id is fn
        sheet.cell(row=next_row, column=1, value=fn)
        sheet.cell(row=next_row, column=2, value=str(short_path))

        title_change_dict = get_title_and_bibl_scope(xml_file)
        if title_change_dict:
            bibl_scope = title_change_dict.get("bibl_scope",  "")
            title = title_change_dict.get("title", "")
        else:
            bibl_scope = ""
            title = ""
        sheet.cell(row=next_row, column=3, value=title)
        sheet.cell(row=next_row, column=4, value=bibl_scope)

        front_file = find_file_recursively(stage_front_path, xml_file.name)
        if front_file:
            title_front_change_dict = get_title_and_bibl_scope(front_file)
        else:
            title_front_change_dict = None

        if title_front_change_dict:
            bibl_scope = title_front_change_dict.get("bibl_scope", "")
            title = title_front_change_dict.get("title", "")
        else:
            bibl_scope = ""
            title = ""

        front_short_path = front_file.relative_to(TOP_DIR)
        sheet.cell(row=next_row, column=5, value=str(front_short_path))
        sheet.cell(row=next_row, column=6, value=title)
        sheet.cell(row=next_row, column=7, value=bibl_scope)

        sheet.cell(row=next_row, column=8, value=str(bibllist_bio_short_path))
        bibl_title_changes_dict = get_bibllist_bio_title_and_id(xml_file)
        if bibl_title_changes_dict:
            bibl_original_title = bibl_title_changes_dict.get("bibl_original_title", "")
            xml_id = bibl_title_changes_dict.get("xml_id", "")
        else:
            bibl_original_title = ""
            xml_id = ""

        sheet.cell(row=next_row, column=9, value=xml_id)
        sheet.cell(row=next_row, column=10, value=bibl_original_title)
        next_row += 1


    workbook.save(export_csv_path / f'{new_file_prefix}{last_git_hash}.xlsx')

    elapsed = time.perf_counter() - start
    logger.info(f"Total time: {elapsed:.2f} seconds, rows: {next_row-1}")


if __name__ == "__main__":
    main()
