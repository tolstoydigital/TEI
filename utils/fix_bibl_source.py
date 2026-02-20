import os
import re
import time
from pathlib import Path
from typing import Optional, Dict, Set

from lxml import etree
import logging
from csv import DictWriter
from openpyxl import load_workbook, Workbook
import subprocess


def setup_logger(name: str = None, level=logging.INFO, logfile: Path | None = None):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.handlers.clear()

    formatter = logging.Formatter("%(levelname)s: %(message)s")  # Без даты

    # Вывод в консоль
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # Если указан файл — пишем ещё и туда
    if logfile:
        file_handler = logging.FileHandler(logfile, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


# Абсолютный путь к текущему файлу
BASE_DIR = Path(__file__).resolve().parent
TOP_DIR = BASE_DIR.parent.parent

SCRIPT_NAME = Path(__file__).stem

bibllist_bio_path = BASE_DIR.parent / "reference" / "bibllist_bio.xml"
bibllist_bio_path = bibllist_bio_path.resolve()
bibllist_bio_short_path = bibllist_bio_path.relative_to(TOP_DIR)

source_list_path = BASE_DIR.parent / "reference" / "sourceList.xml"
source_list_path = source_list_path.resolve()


SD = "-"
LD = "—"
BIO_C = "С."
pattern = r"(?<=\b\d{4})\s*[-–—]\s*(?=\d{4}\b)"


logger = setup_logger(__name__, logfile=Path(f"{SCRIPT_NAME}.log"))


def get_source_list_set(filename: Path = source_list_path) -> Set[str]:

    tree = etree.parse(filename)
    root = tree.getroot()

    ns = root.nsmap.get(None)
    nsmap = (
        {"tei": ns, "xml": "http://www.w3.org/XML/1998/namespace"}
        if ns
        else {"xml": "http://www.w3.org/XML/1998/namespace"}
    )
    # <item n="1" xml:id="Azbuka_1873">...
    out = set()
    for item in root.findall(".//tei:item[@xml:id]", namespaces=nsmap):
        xml_id = item.get("{http://www.w3.org/XML/1998/namespace}id")
        if xml_id:  # защита от пустых атрибутов
            out.add(xml_id)

    return out

def get_bio_set(filename: Path = bibllist_bio_path) -> Set[str]:
    tree = etree.parse(filename)
    root = tree.getroot()

    ns = root.nsmap.get(None)
    nsmap = {
        "tei": ns,
        "xml": "http://www.w3.org/XML/1998/namespace"
    } if ns else {"xml": "http://www.w3.org/XML/1998/namespace"}

    out = set()
    # собираем все <ref> внутри <relatedItem>, у которых есть xml:id
    for ref in root.findall(".//tei:relatedItem/tei:ref[@xml:id]", namespaces=nsmap):
        xml_id = ref.get("{http://www.w3.org/XML/1998/namespace}id")
        if xml_id:
            out.add(xml_id)

    return out

def get_slovo_set() -> Set[str]:
    ref_path = (
            BASE_DIR.parent
            / "utils"
            / "doc"
            / "post-evaluation-table - True_False & pages Review_Edits_02.xlsx"
    )
    ref_path = ref_path.resolve()

    try:
        workbook = load_workbook(ref_path)
    except FileNotFoundError:
        print(f"Файл {ref_path} не найден.")
        return

    # sheet = workbook.active
    sheet = workbook["Sheet1"]

    headers = [cell.value for cell in sheet[1]]

    out = set()
    for row in sheet.iter_rows(
            min_row=2,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
    ):
        if not any(row):
            continue
        data = dict(zip(headers, row))

        td_source_item_id = data.get("TD_source_item_ID", "")  # N
        source = data.get("source", "")  # T
        if source == 'slovo-tolstogo' and td_source_item_id:
            out.add(td_source_item_id)

    return out

def get_title_and_bibl_scope(
    xml_path: Path, save_to: Optional[Path] = None
) -> Dict[str, str]:
    short_path = xml_path.relative_to(TOP_DIR)

    tree = etree.parse(xml_path)
    root = tree.getroot()

    ns = root.nsmap.get(None)
    nsmap = (
        {"tei": ns, "xml": "http://www.w3.org/XML/1998/namespace"}
        if ns
        else {"xml": "http://www.w3.org/XML/1998/namespace"}
    )

    for title_el in root.findall(
        ".//tei:title[@type='bibl']" if ns else ".//title[@type='bibl']",
        namespaces=nsmap,
    ):
        # original_title = title_el.text
        if title_el.text is None:
            # continue  # защита от пустых тегов
            source_old_title_text = ""
            old_title_text = ""
        else:
            source_old_title_text = title_el.text.strip()
            old_title_text = (
                source_old_title_text  # replace_title(source_old_title_text)
            )

        bibl_scope_text = ""
        for bibl_scope_el in root.findall(
            ".//tei:biblScope[@unit='page']" if ns else ".//biblScope[@unit='page']",
            namespaces=nsmap,
        ):
            bibl_scope_text = bibl_scope_el.text.strip()

        return {"bibl_scope": bibl_scope_text, "title": old_title_text}

    return None


def get_bibllist_bio_title_and_id(filename: Path) -> Optional[Dict[str, str]]:

    xml_id = filename.stem
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }
    xpath = f".//tei:relatedItem[tei:ref[@xml:id=\"{xml_id}\"]]/tei:title[@type='bibl']"
    titles = bib_root.xpath(xpath, namespaces=ns)
    for t in titles:
        # logger.info(f"====================================== {xml_id}")
        # logger.info(f"BIBLIO <<< {t.text}")
        # logger.info(f"BIBLIO >>> {new_title}")
        # t.text = new_title
        bib_changes = {"bibl_original_title": t.text, "xml_id": xml_id}
        # bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
        return bib_changes
    return None




def fix_in_bibllist_bio_case_1():
    bib_tree = etree.parse(bibllist_bio_path)
    bib_root = bib_tree.getroot()

    ns = {
        "tei": "http://www.tei-c.org/ns/1.0",
        "xml": "http://www.w3.org/XML/1998/namespace",
    }

    xpath = (f'.//tei:relation[@source and @type="source"]')  # /tei:biblpoint

    xp_title = f".//tei:title[@type='main']"
    xp_date = f".//tei:date[@type='editor']"

    rel_it_nodes = bib_root.xpath(xpath, namespaces=ns)
    logger.info(f"======= CASE_1")

    sl = get_source_list_set()
    # print(len(sl))
    bio = get_bio_set()
    # print(len(bio))
    slovo = get_slovo_set()
    # print(len(slovo))

    bio_count = 0
    slovo_count = 0
    sl_count = 0
    if rel_it_nodes:
        for _rel in rel_it_nodes:
            ref = _rel.get("ref")
            source = _rel.get("source")
            # logger.info(
            #     f'      ------ {etree.tostring(_rel, pretty_print=True, encoding="unicode")}\n'
            #     f'{ref=} {source=}'
            # )

            some = False
            if ref in slovo and source != 'slovo-tolstogo':
                logger.info(f'      WRONG source {source} should be slovo-tolstogo {etree.tostring(_rel, pretty_print=True, encoding="unicode")}')
                _rel.set("source", 'slovo-tolstogo')
                some = True
                slovo_count += 1

            if ref in bio and source != 'tolstoy-bio':
                logger.info(
                    f'      WRONG source {source} should be tolstoy-bio {etree.tostring(_rel, pretty_print=True, encoding="unicode")}')
                _rel.set("source", 'tolstoy-bio')
                some = True
                bio_count += 1

            if ref in sl and source != 'source-list':
                logger.info(f'      WRONG source {source} should be source-list {etree.tostring(_rel, pretty_print=True, encoding="unicode")}')
                _rel.set("source", 'source-list')
                some = True
                sl_count += 1

            if some:
                logger.info('='*50)


    else:
        logger.info(f"      NO relatedItem")

    logger.info(f">>>> source_list errors: {sl_count}, slovo_tolstogo errors: {slovo_count}, biblist_bio errors: {bio_count}")
    bib_tree.write(bibllist_bio_path, encoding="UTF-8", xml_declaration=True, pretty_print=True)
    return


def main():
    start = time.perf_counter()

    fix_in_bibllist_bio_case_1()

    elapsed = time.perf_counter() - start
    logger.info(f"Total time: {elapsed:.2f} seconds")


if __name__ == "__main__":
    main()
