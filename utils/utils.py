import logging
import os
import subprocess
from pathlib import Path
from typing import Iterator, Dict, Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPO_PATH = Path('..')


def convert_language_to_iso(language: str) -> str:
    languages = {
        'english': 'en',
        'french': 'fr',
        '': 'ru'
    }
    return languages[language] if language in languages else language


def read_xml(xml, mode='r'):
    with open(xml, mode) as file:
        return file.read()


def setup_logger(name: str = None, level=logging.INFO, logfile: Path | None = None):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.handlers.clear()

    formatter = logging.Formatter("%(levelname)s: %(message)s")  # Без даты

    # Вывод в консоль
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # Если указан файл — пишем ещё и туда
    if logfile:
        file_handler = logging.FileHandler(logfile, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


def iter_xml_files(root_dir):
    root_dir = Path(root_dir)
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            if filename.lower().endswith(".xml"):
                yield Path(dirpath) / filename


def iter_xlsx_rows_as_column_letter_dict(
    file_path: str,
    sheet_name: str,
    start_row: int = 1,
) -> Iterator[Dict[str, Any]]:
    """
    Итерация по строкам xlsx-файла с остановкой на первой пустой строке.
    Возвращает словарь вида {'A': value, 'B': value, ...}

    :param file_path: путь к xlsx-файлу
    :param sheet_name: имя листа
    :param start_row: строка начала (1-based)
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    max_col = ws.max_column

    for row in ws.iter_rows(
        min_row=start_row,
        max_col=max_col,
        values_only=True,
    ):
        # полностью пустая строка → останавливаемся
        if all(cell is None for cell in row):
            break

        yield {
            get_column_letter(col_idx + 1): value for col_idx, value in enumerate(row)
        }


def git_short_sha() -> str:
    return (
        subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
        or ""
    )


def find_file_recursively(search_directory: str, filename: str) -> Optional[str]:
    """
    Ищет файл filename в папке search_directory и всех её подпапках.

    Args:
        search_directory (str): Путь к папке, откуда начинать поиск.
        filename (str): Точное имя файла (например, 'data.xml').

    Returns:
        Optional[str]: Полный абсолютный путь к файлу (str) или None, если файл не найден.
    """
    path_obj = Path(search_directory)

    # Проверяем, существует ли вообще папка поиска
    if not path_obj.exists():
        return None

    try:
        # rglob('*') ищет рекурсивно.
        # Если мы передаем конкретное имя файла в rglob, он будет искать именно его.
        # next() берет первое совпадение.
        found_file = next(path_obj.rglob(filename))

        # .resolve() возвращает абсолютный путь (например, /home/user/projects/...)
        return found_file.resolve()

    except StopIteration:
        # Если генератор пуст (файл не найден), next() вызовет StopIteration
        return None
